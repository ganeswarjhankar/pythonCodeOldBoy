import openpyxl

def getRowCount(file,sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def readData(file,sheetName,rowno,columnno):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    return (sheet.cell(row=rowno,column=columnno).value

def writeData(file, sheetName, rowno, columnno,data):
        workBook = openpyxl.load_workbook(file)
        sheet = workBook.get_sheet_by_name(sheetName)
        sheet.cell(row=rowno, column=columnno).value = data
        workBook.save(file)

